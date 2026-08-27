"""
PDF & Image to Excel Converter — Backend API
---------------------------------------------
Accepts a PDF or Image (.png, .jpg, .jpeg, .webp, .bmp) upload,
extracts tables using pdfplumber, PyMuPDF (for scanned PDFs), & Tesseract OCR,
validates financial arithmetic (Prev Balance + Credit - Debit = Balance),
and returns a downloadable styled .xlsx file with summary reconciliation dashboard.
"""
import io
import os
import re
import pdfplumber
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS

import pytesseract
import cv2
from PIL import Image

try:
    import pymupdf as fitz  # PyMuPDF for scanned PDF page rendering
except ImportError:
    try:
        import fitz
    except ImportError:
        fitz = None

# Configure Tesseract binary path on Windows
tess_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
if os.path.exists(tess_path):
    pytesseract.pytesseract.tesseract_cmd = tess_path

app = Flask(__name__)
CORS(app)

MAX_FILE_SIZE_MB = 20
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_SIZE_MB * 1024 * 1024

IMAGE_EXTENSIONS = ('.png', '.jpg', '.jpeg', '.webp', '.bmp', '.tiff')


class NamedBytesIO(io.BytesIO):
    """BytesIO wrapper with a .filename attribute for seamless image list handling."""
    def __init__(self, initial_bytes=b"", filename="page.png"):
        super().__init__(initial_bytes)
        self.filename = filename


def _clean_header(header):
    """Clean raw header row: remove None/empty columns and deduplicate names."""
    valid_indices = []
    cleaned = []
    seen = {}
    for i, h in enumerate(header):
        name = (h or "").strip()
        if not name:
            continue
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        cleaned.append(name)
        valid_indices.append(i)
    return cleaned, valid_indices


def _is_empty_row(row):
    """Check if a row is entirely empty/None."""
    return all((cell is None or str(cell).strip() == "") for cell in row)


def _clean_dataframe(df):
    """Post-process DataFrame."""
    df = df.dropna(how="all").reset_index(drop=True)
    mask = df.apply(lambda row: not all(
        str(v).strip() == "" or (isinstance(v, float) and np.isnan(v))
        for v in row
    ), axis=1)
    df = df[mask].reset_index(drop=True)

    for col in df.columns:
        df[col] = df[col].apply(lambda x:
            str(x).replace("\n", " ").strip()
            if pd.notna(x) and str(x).strip() != "" else ""
        )
    return df


def _calculate_missing_ratio(df):
    """Calculate what fraction of cells are empty/missing."""
    if df.empty or len(df.columns) <= 2:
        return 1.0
    data_cols = df.iloc[:, 2:]
    total = data_cols.size
    if total == 0:
        return 1.0
    empty_count = data_cols.apply(
        lambda col: col.apply(lambda x: pd.isna(x) or str(x).strip() == "")
    ).sum().sum()
    return empty_count / total


def _parse_text_lines_to_df(lines, p1_text="", p_last_text=""):
    """
    Core text line parser shared by PDF text extraction, Scanned PDF OCR, and Image OCR.
    Extracts metadata, dates, descriptions, credit, debit, available balance,
    and performs financial arithmetic reconciliation validation.
    """
    date_pattern = re.compile(r'^(?:\d{2}[/\-\s](?:\d{2}|\w{3})[/\-\s]\d{4}|\d{4}[/\-\s]\d{4})\b')
    amount_pattern = re.compile(r'(?:[+-]\s*PKR\s*[\d,.]+|PKR\s*[\d,.]+|(?<=\s)[\d,]+\.\d{2}(?=\s|$)|(?<=\s)[\d,]+\.\d{1,2}(?=\s|$))')
    footer_pattern = re.compile(r'^\d+\s+\d{2}\s+\w{3}\s+\d{4},\s+\d{2}:\d{2}$|^\d+\s*\|\s*Page$|^\*\*\*\*\*\*End of statement\*\*\*\*\*\*')

    all_entries = []
    total_debit = 0.0
    total_credit = 0.0
    current_entry = None

    for line in lines:
        line = line.strip()
        if not line or footer_pattern.match(line) or 'This is a system generated report' in line or 'STATEMENT PERIOD' in line or 'Balance B/F' in line:
            continue

        # Clean OCR noise ($/S/s before comma or amount)
        line = line.replace('$,', ' ').replace('S,', ' ').replace('s,', ' ')

        # Normalize OCR number misreads (e.g. 6.067.389.60 -> 6,067,389.60 and 5.560,389.60 -> 5,560,389.60)
        line = re.sub(r'(\d{1,3})\.(\d{3})[\.,](\d{3})\.(\d{2})', r'\1,\2,\3.\4', line)
        line = re.sub(r'(\d{1,3})\.(\d{3})[\.,](\d{2})', r'\1,\2.\3', line)

        # Normalize missing slash in OCR dates (e.g. 0201/2026 -> 02/01/2026)
        norm_line = re.sub(r'^(\d{2})(\d{2})/(\d{4})', r'\1/\2/\3', line)

        if date_pattern.match(norm_line):
            if current_entry:
                all_entries.append(current_entry)

            m = re.match(r'^(\d{2}[/\-\s](?:\d{2}|\w{3})[/\-\s]\d{4})(?:\s+(\d{2}[/\-\s](?:\d{2}|\w{3})[/\-\s]\d{4}))?\s+(.*)', norm_line)
            if m:
                tx_date = m.group(1)
                rest = m.group(3)

                amounts = amount_pattern.findall(rest)
                desc = amount_pattern.sub('', rest).strip()
                desc = re.sub(r'\s{2,}', ' ', desc).strip()

                current_entry = {
                    'date': tx_date,
                    'desc': desc,
                    'amounts': amounts,
                }
        elif current_entry:
            if re.match(r'^\d+\s+\d{2}\s+\w{3}\s+\d{4}', line):
                continue

            amounts_in_line = amount_pattern.findall(line)
            desc_part = amount_pattern.sub('', line).strip()
            desc_part = re.sub(r'\s{2,}', ' ', desc_part).strip()

            if desc_part and desc_part not in (
                "Account Statement", "Booking Date", "Description",
                "Credit", "Debit", "Available Balance", "Balance",
                "Date Description Debit Credit Balance",
                "Booking Date Description Credit Debit Available Balance",
                "Transaction Value Instrument Cr. Remaining",
                "Nature of Transaction Dr. Amount",
                "Date Date Number Amount Balance",
            ) and "of this statement, otherwise" not in desc_part:
                current_entry['desc'] += ' ' + desc_part
            current_entry['amounts'].extend(amounts_in_line)

    if current_entry:
        all_entries.append(current_entry)

    if not all_entries:
        return []

    def _amount_to_number(amt_str):
        if not amt_str or str(amt_str).strip() == "":
            return None
        clean = amt_str.replace('+', '').replace('-', '').replace('PKR', '').replace(',', '').replace(' ', '').strip()
        clean = clean.replace('$0', '50').replace('S0', '50').replace('s0', '50').replace('$', '5')
        if clean.count('.') > 1:
            parts = clean.split('.')
            clean = ''.join(parts[:-1]) + '.' + parts[-1]
        try:
            val = float(clean)
            if val == 0.0:
                return None
            return val
        except ValueError:
            return None

    # Parse metadata header
    account_title = ""
    account_number = ""
    iban = ""
    currency = "Pakistan Rupee (PKR)"
    from_date = ""
    to_date = ""
    opening_balance = None
    closing_balance = None
    bank_name = "Bank"

    lines_p1 = [l.strip() for l in p1_text.split('\n') if l.strip()]
    lines_last = [l.strip() for l in p_last_text.split('\n') if l.strip()]

    for i in range(len(lines_p1)):
        line = lines_p1[i]

        # Multi-column Opening/Closing Balance on adjacent lines
        if 'Opening Balance' in line and 'Closing Balance' in line:
            if i + 1 < len(lines_p1):
                val_line = lines_p1[i + 1]
                amts = re.findall(r'[\d,]+\.\d{2}', val_line)
                if len(amts) >= 2:
                    opening_balance = _amount_to_number(amts[0])
                    closing_balance = _amount_to_number(amts[1])
                elif len(amts) == 1:
                    opening_balance = _amount_to_number(amts[0])

        if ('ACCOUNT NUMBER' in line or 'Account Number' in line) and not account_number:
            m_acc = re.search(r'(.*?)\s+ACCOUNT NUMBER\s+.*?\s*(\d{10,})', line, re.I)
            if m_acc:
                account_title = m_acc.group(1).strip()
                account_number = m_acc.group(2).strip()

        if 'Account Title' in line and 'Account Number' in line:
            if i + 1 < len(lines_p1):
                val_line = lines_p1[i + 1]
                iban_m = re.search(r'(PK\d{2}[A-Z]{4}\d+)', val_line)
                if iban_m:
                    iban = iban_m.group(1)
                    before_iban = val_line[:val_line.find(iban)].strip()
                    acc_num_m = re.search(r'(\d{10,16})', before_iban)
                    if acc_num_m:
                        account_number = acc_num_m.group(1)
                        account_title = before_iban[:before_iban.find(account_number)].strip()
                    else:
                        account_title = before_iban

        if 'Account Number:' in line or 'Account Title:' in line or 'Opening Balance:' in line or 'Closing Balance:' in line:
            if 'Account Title:' in line:
                account_title = line.split('Account Title:', 1)[-1].strip()
            if 'Account Number:' in line:
                account_number = line.split('Account Number:', 1)[-1].strip()
            if 'Currency:' in line:
                currency = line.split('Currency:', 1)[-1].strip()
            if 'Opening Balance:' in line:
                opening_balance = _amount_to_number(line.split('Opening Balance:', 1)[-1].strip())
            if 'Closing Balance:' in line:
                closing_balance = _amount_to_number(line.split('Closing Balance:', 1)[-1].strip())

        if ('IBAN' in line or 'wan' in line.lower()) and not iban:
            m_iban = re.search(r'(PK\d{2}[A-Z0-9]{18,22})', line, re.I)
            if m_iban:
                iban = m_iban.group(1).strip()

        if re.search(r'\d{2}/\d{2}/\d{4}.*?TO.*?\d{2}/\d{2}/\d{4}', line, re.I):
            m_per = re.search(r'(\d{2}/\d{2}/\d{4})\s*.*?TO.*?\s*(\d{2}/\d{2}/\d{4})', line, re.I)
            if m_per:
                from_date = m_per.group(1).strip()
                to_date = m_per.group(2).strip()
        elif not from_date and re.search(r'\d{2}/\d{2}/\d{4}', line):
            dates_in_line = re.findall(r'(\d{2}/\d{2}/\d{4})', line)
            if len(dates_in_line) >= 2:
                from_date = dates_in_line[0]
                to_date = dates_in_line[1]

        if ('Balance B/F' in line or 'Balance BF' in line or 'balance b/f' in line.lower()) and opening_balance is None:
            m_bal = re.search(r'[\d,]+\.\d{2}', line)
            if m_bal:
                opening_balance = _amount_to_number(m_bal.group(0))
            else:
                idx = lines_p1.index(line) if line in lines_p1 else -1
                if idx != -1 and idx + 1 < len(lines_p1):
                    opening_balance = _amount_to_number(lines_p1[idx + 1])

    if closing_balance is None:
        for line in lines_last:
            if 'Closing Balance' in line:
                m_cb = re.findall(r'[\d,]+\.\d{2}', line)
                if m_cb:
                    closing_balance = _amount_to_number(m_cb[-1])

    full_p1_text = ' '.join(lines_p1)
    if "MEZN" in (iban or "") or "Meezan" in full_p1_text:
        bank_name = "Meezan Bank"
    elif "BPUN" in (iban or "").upper() or "PUNJAB" in full_p1_text.upper() or "BOP" in full_p1_text.upper():
        bank_name = "The Bank of Punjab"
    elif "ALLIED" in full_p1_text.upper() or "ABL" in full_p1_text.upper() or "0010" in account_number or "0010" in full_p1_text:
        bank_name = "Allied Bank"
    else:
        bank_name = "Bank"

    statement_period = ""
    if from_date and to_date:
        statement_period = f"{from_date} to {to_date}"
    elif from_date:
        statement_period = from_date

    records = []
    header_rows = []

    header_rows.append([f"{bank_name} Account Statement", "", "", "", "", ""])
    header_rows.append(["", "", "", "", "", ""])
    header_rows.append(["Account Title:", account_title, "", "", "", ""])
    header_rows.append(["Account Number:", account_number, "", "", "", ""])
    header_rows.append(["IBAN:", iban, "", "", "", ""])
    header_rows.append(["Currency:", currency, "", "", "", ""])
    header_rows.append(["Statement Period:", statement_period, "", "", "", ""])
    header_rows.append(["Opening Balance:", opening_balance if opening_balance is not None else "", "", "", "", ""])
    header_rows.append(["Closing Balance:", closing_balance if closing_balance is not None else "", "", "", "", ""])
    header_rows.append(["", "", "", "", "", ""])
    header_rows.append(["Booking Date", "Description", "Credit", "Debit", "Available Balance", "Validation Status"])

    running_balance = opening_balance
    review_needed_count = 0

    for e in all_entries:
        if not e['amounts']:
            continue

        credit = None
        debit = None
        balance = None

        valid_amounts = e['amounts']
        has_prefix = any(('PKR' in a or '+' in a or '-' in a) for a in valid_amounts)

        if has_prefix:
            for amt in valid_amounts:
                amt_clean = amt.strip()
                if amt_clean.startswith('+'):
                    credit = _amount_to_number(amt_clean)
                    if credit: total_credit += credit
                elif amt_clean.startswith('-'):
                    debit = _amount_to_number(amt_clean)
                    if debit: total_debit += debit
                elif 'PKR' in amt_clean:
                    balance = _amount_to_number(amt_clean)
        else:
            clean_nums = [_amount_to_number(a) for a in valid_amounts]
            valid_clean_nums = [n for n in clean_nums if n is not None]

            if len(clean_nums) >= 3:
                debit = clean_nums[0]
                credit = clean_nums[1]
                balance = clean_nums[2]
                if debit: total_debit += debit
                if credit: total_credit += credit
            elif len(valid_clean_nums) >= 2:
                tx_amt = valid_clean_nums[0]
                balance = valid_clean_nums[1]
                desc_lower = e['desc'].lower()
                if any(kw in desc_lower for kw in ['credit', 'encashment', 'deposit', 'received', 'refund', 'rev', 'rtgs']):
                    credit = tx_amt
                    if credit: total_credit += credit
                else:
                    debit = tx_amt
                    if debit: total_debit += debit
            elif len(valid_clean_nums) == 1:
                balance = valid_clean_nums[0]

        desc_clean = re.sub(r'\s{2,}', ' ', e['desc'].strip()).strip()

        # Financial Arithmetic Reconciliation Validation
        validation_status = "✅ Verified"
        if running_balance is not None and balance is not None:
            expected_balance = running_balance + (credit or 0.0) - (debit or 0.0)
            if abs(expected_balance - balance) > 0.10:
                validation_status = "⚠️ Review Required"
                review_needed_count += 1
            running_balance = balance
        elif balance is not None:
            running_balance = balance

        records.append({
            "date": e['date'],
            "desc": desc_clean,
            "credit": credit,
            "debit": debit,
            "balance": balance,
            "status": validation_status
        })

        c_str = f"{credit:.2f}" if credit is not None else ""
        d_str = f"{debit:.2f}" if debit is not None else ""
        b_str = f"{balance:.2f}" if balance is not None else ""
        header_rows.append([e['date'], desc_clean, c_str, d_str, b_str, validation_status])

    header_rows.append(["", "", "", "", "", ""])
    header_rows.append(["", "Total", f"{total_credit:.2f}", f"{total_debit:.2f}", "", ""])

    col_names = [f"{bank_name} Account Statement" if i == 0 else f"Unnamed: {i}" for i in range(6)]
    df = pd.DataFrame(header_rows, columns=col_names)
    df.attrs["sheet_name"] = "Account Statement"
    df.attrs["has_header"] = False
    df.attrs["is_statement"] = True
    df.attrs["bank_name"] = bank_name
    df.attrs["metadata"] = {
        "account_title": account_title,
        "account_number": account_number,
        "iban": iban,
        "currency": currency,
        "statement_period": statement_period,
        "opening_balance": opening_balance,
        "closing_balance": closing_balance,
        "review_count": review_needed_count,
        "total_records": len(records)
    }
    df.attrs["records"] = records
    return [df]


def _extract_tables_via_text(file_stream):
    """Fallback text-based extraction for PDFs."""
    lines = []
    p1_text = ""
    p_last_text = ""

    file_stream.seek(0)
    with pdfplumber.open(file_stream) as pdf:
        if pdf.pages:
            p1_text = pdf.pages[0].extract_text() or ""
            p_last_text = pdf.pages[-1].extract_text() or "" if len(pdf.pages) > 1 else p1_text
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    lines.extend(t.split('\n'))

    return _parse_text_lines_to_df(lines, p1_text, p_last_text)


def _extract_tables_from_scanned_pdf(file_stream):
    """
    Scanned PDF Fallback Engine:
    Renders PDF pages to high-DPI (300 DPI) images using PyMuPDF (fitz)
    and passes them to the OpenCV + Tesseract OCR engine.
    """
    if not fitz:
        return []

    file_stream.seek(0)
    file_bytes = file_stream.read()
    doc = fitz.open(stream=file_bytes, filetype="pdf")

    image_files = []
    for p_idx, page in enumerate(doc, start=1):
        pix = page.get_pixmap(dpi=300)
        img_bytes = pix.tobytes("png")
        image_files.append(NamedBytesIO(img_bytes, filename=f"page_{p_idx:03d}.png"))

    if not image_files:
        return []

    return _extract_tables_from_images(image_files)


def _extract_tables_from_images(image_files):
    """
    Extract tables and text from one or multiple uploaded image files using
    OpenCV unsharp-mask sharpening, OTSU thresholding, and Tesseract OCR.
    """
    combined_lines = []
    first_ocr_text = ""
    last_ocr_text = ""

    sorted_files = sorted(image_files, key=lambda f: getattr(f, 'filename', 'page.png'))

    for idx, img_file in enumerate(sorted_files):
        img_file.seek(0)
        file_bytes = np.frombuffer(img_file.read(), np.uint8)
        img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)

        if img is None:
            continue

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        h, w = gray.shape[:2]
        if w < 1600:
            scale = 1600 / float(w)
            gray = cv2.resize(gray, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_CUBIC)

        gaussian = cv2.GaussianBlur(gray, (0, 0), 3)
        sharpened = cv2.addWeighted(gray, 1.5, gaussian, -0.5, 0)
        _, thresh = cv2.threshold(sharpened, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        ocr_text = pytesseract.image_to_string(thresh, config='--psm 6')
        if not ocr_text.strip():
            ocr_text = pytesseract.image_to_string(gray, config='--psm 6')
        if not ocr_text.strip():
            ocr_text = pytesseract.image_to_string(thresh)

        if idx == 0:
            first_ocr_text = ocr_text
        last_ocr_text = ocr_text

        lines = [l.strip() for l in ocr_text.split('\n') if l.strip()]
        combined_lines.extend(lines)

    if not combined_lines:
        return []

    return _parse_text_lines_to_df(combined_lines, first_ocr_text, last_ocr_text)


def extract_tables_from_pdf(file_stream):
    """
    Reads PDF file stream and returns list of DataFrames.
    Automatically detects scanned PDFs and falls back to High-DPI OCR.
    """
    raw_tables = []

    with pdfplumber.open(file_stream) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            for table_index, table in enumerate(tables, start=1):
                if not table or len(table) < 1:
                    continue

                raw_header = table[0]
                raw_rows = table[1:]

                cleaned_header, valid_indices = _clean_header(raw_header)
                if not cleaned_header:
                    continue

                cleaned_rows = []
                for row in raw_rows:
                    if _is_empty_row(row):
                        continue
                    cleaned_row = []
                    for idx in valid_indices:
                        if idx < len(row):
                            cleaned_row.append(row[idx])
                        else:
                            cleaned_row.append("")
                    if _is_empty_row(cleaned_row):
                        continue
                    cleaned_rows.append(cleaned_row)

                if not cleaned_rows:
                    continue

                df = pd.DataFrame(cleaned_rows, columns=cleaned_header)
                df = _clean_dataframe(df)

                if df.empty:
                    continue

                raw_tables.append((
                    tuple(cleaned_header),
                    df,
                    page_number,
                    table_index
                ))

    merged = []
    for header_key, df, page_num, tbl_idx in raw_tables:
        if merged and merged[-1][0] == header_key:
            merged[-1] = (
                header_key,
                pd.concat([merged[-1][1], df], ignore_index=True),
                merged[-1][2],
                merged[-1][3],
            )
        else:
            merged.append((header_key, df, page_num, tbl_idx))

    dataframes = []
    for i, (header_key, df, page_num, tbl_idx) in enumerate(merged):
        df = _clean_dataframe(df)
        if df.empty:
            continue
        df.attrs["sheet_name"] = f"Page{page_num}_Table{tbl_idx}"
        dataframes.append(df)

    if not dataframes:
        text_dfs = _extract_tables_via_text(file_stream)
        if text_dfs and len(text_dfs[0]) > 0:
            return text_dfs
        # Scanned PDF Fallback
        return _extract_tables_from_scanned_pdf(file_stream)

    if dataframes:
        main_df = max(dataframes, key=lambda d: len(d))
        missing_ratio = _calculate_missing_ratio(main_df)
        if missing_ratio > 0.5:
            text_dfs = _extract_tables_via_text(file_stream)
            if text_dfs and len(text_dfs[0]) > 0:
                return text_dfs
            return _extract_tables_from_scanned_pdf(file_stream)

    return dataframes


def _build_styled_excel_file(dataframes):
    """
    Builds a beautifully styled openpyxl Excel file with:
    - Executive Header & Reconciliation Dashboard
    - Soft yellow highlight for rows flagged as ⚠️ Review Required
    - Auto Column Widths & Gridlines
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    warn_font = Font(name="Calibri", size=11, bold=True, color="7F6000")
    pass_font = Font(name="Calibri", size=11, bold=True, color="375623")

    total_border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    for df in dataframes:
        sheet_name = df.attrs.get("sheet_name", "Sheet1")[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        is_statement = df.attrs.get("is_statement", False)

        if is_statement:
            meta = df.attrs.get("metadata", {})
            bank_name = df.attrs.get("bank_name", "Bank")
            review_count = meta.get("review_count", 0)

            ws.cell(row=1, column=1, value=f"{bank_name} Account Statement").font = title_font

            rec_status = "✅ 100% Verified (0 Errors)" if review_count == 0 else f"⚠️ {review_count} Rows Need Review"

            meta_items = [
                ("Account Title:", meta.get("account_title", "")),
                ("Account Number:", str(meta.get("account_number", ""))),
                ("IBAN:", str(meta.get("iban", ""))),
                ("Currency:", meta.get("currency", "Pakistan Rupee (PKR)")),
                ("Statement Period:", meta.get("statement_period", "")),
                ("Opening Balance:", meta.get("opening_balance", None)),
                ("Closing Balance:", meta.get("closing_balance", None)),
                ("Reconciliation Status:", rec_status),
            ]

            curr_row = 3
            for label, val in meta_items:
                lbl_cell = ws.cell(row=curr_row, column=1, value=label)
                lbl_cell.font = bold_font
                val_cell = ws.cell(row=curr_row, column=2)

                if "Balance" in label and val is not None and str(val) != "":
                    try:
                        val_cell.value = float(val)
                        val_cell.number_format = "#,##0.00"
                        val_cell.alignment = Alignment(horizontal="right")
                    except ValueError:
                        val_cell.value = str(val)
                elif label == "Reconciliation Status:":
                    val_cell.value = str(val)
                    val_cell.font = pass_font if "✅" in str(val) else warn_font
                    val_cell.fill = green_fill if "✅" in str(val) else yellow_fill
                else:
                    val_cell.value = str(val) if val is not None else ""
                    val_cell.number_format = "@"
                    val_cell.alignment = Alignment(horizontal="left")
                    val_cell.font = regular_font
                curr_row += 1

            curr_row = 11

            headers = ["Booking Date", "Description", "Credit", "Debit", "Available Balance", "Validation Status"]
            for col_idx, h_text in enumerate(headers, 1):
                c = ws.cell(row=curr_row, column=col_idx, value=h_text)
                c.fill = navy_fill
                c.font = white_bold_font
                if h_text in ["Credit", "Debit", "Available Balance"]:
                    c.alignment = Alignment(horizontal="right", vertical="center")
                elif h_text in ["Booking Date", "Validation Status"]:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

            first_data_row = 12
            records = df.attrs.get("records", [])
            row_idx = first_data_row

            for r in records:
                is_warn = "Review Required" in r.get("status", "")
                row_fill = yellow_fill if is_warn else None

                c_date = ws.cell(row=row_idx, column=1, value=r.get("date", ""))
                c_date.alignment = Alignment(horizontal="center")
                c_date.font = regular_font
                if row_fill: c_date.fill = row_fill

                c_desc = ws.cell(row=row_idx, column=2, value=r.get("desc", ""))
                c_desc.alignment = Alignment(horizontal="left", wrap_text=True)
                c_desc.font = regular_font
                if row_fill: c_desc.fill = row_fill

                c_cred = ws.cell(row=row_idx, column=3)
                if r.get("credit") is not None:
                    try:
                        c_cred.value = float(r["credit"])
                        c_cred.number_format = "#,##0.00"
                    except ValueError:
                        c_cred.value = str(r["credit"])
                c_cred.alignment = Alignment(horizontal="right")
                c_cred.font = regular_font
                if row_fill: c_cred.fill = row_fill

                c_deb = ws.cell(row=row_idx, column=4)
                if r.get("debit") is not None:
                    try:
                        c_deb.value = float(r["debit"])
                        c_deb.number_format = "#,##0.00"
                    except ValueError:
                        c_deb.value = str(r["debit"])
                c_deb.alignment = Alignment(horizontal="right")
                c_deb.font = regular_font
                if row_fill: c_deb.fill = row_fill

                c_bal = ws.cell(row=row_idx, column=5)
                if r.get("balance") is not None:
                    try:
                        c_bal.value = float(r["balance"])
                        c_bal.number_format = "#,##0.00"
                    except ValueError:
                        c_bal.value = str(r["balance"])
                c_bal.alignment = Alignment(horizontal="right")
                c_bal.font = regular_font
                if row_fill: c_bal.fill = row_fill

                c_stat = ws.cell(row=row_idx, column=6, value=r.get("status", "✅ Verified"))
                c_stat.alignment = Alignment(horizontal="center")
                c_stat.font = warn_font if is_warn else pass_font
                if row_fill: c_stat.fill = row_fill

                row_idx += 1

            last_data_row = row_idx - 1

            # Total Row Formulas
            tot_row = row_idx + 1
            lbl_tot = ws.cell(row=tot_row, column=2, value="Total")
            lbl_tot.font = bold_font
            lbl_tot.alignment = Alignment(horizontal="right")

            c_tot_cred = ws.cell(row=tot_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})")
            c_tot_cred.font = bold_font
            c_tot_cred.number_format = "#,##0.00"
            c_tot_cred.border = total_border
            c_tot_cred.alignment = Alignment(horizontal="right")

            c_tot_deb = ws.cell(row=tot_row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})")
            c_tot_deb.font = bold_font
            c_tot_deb.number_format = "#,##0.00"
            c_tot_deb.border = total_border
            c_tot_deb.alignment = Alignment(horizontal="right")

        else:
            write_header = df.attrs.get("has_header", True)
            start_row = 1
            if write_header and list(df.columns):
                for c_idx, col_name in enumerate(df.columns, 1):
                    c = ws.cell(row=1, column=c_idx, value=str(col_name))
                    c.fill = navy_fill
                    c.font = white_bold_font
                start_row = 2

            for r_offset, row in enumerate(df.values):
                r_idx = start_row + r_offset
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    if pd.isna(val) or val is None or str(val).strip() == "":
                        cell.value = ""
                    else:
                        val_str = str(val).strip()
                        try:
                            if re.match(r'^-?[\d,]+\.?\d*$', val_str) and not re.match(r'^0\d{9,}$', val_str):
                                fval = float(val_str.replace(',', ''))
                                cell.value = fval
                                cell.number_format = "#,##0.00"
                                cell.alignment = Alignment(horizontal="right")
                            else:
                                cell.value = val_str
                        except ValueError:
                            cell.value = val_str

        ws.views.sheetView[0].showGridLines = True
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if cell.number_format == "#,##0.00" and isinstance(cell.value, (int, float)):
                    val_str = f"{cell.value:,.2f}"
                max_len = max(max_len, len(val_str))

            if col_letter == 'B':
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 30), 60)
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/", methods=["GET"])
def index():
    candidates = [
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend", "index.html")),
        os.path.abspath(os.path.join(os.path.dirname(__file__), "frontend", "index.html")),
        os.path.abspath("frontend/index.html"),
        os.path.abspath("../frontend/index.html"),
    ]
    for path in candidates:
        if os.path.exists(path):
            return send_file(path)
    return jsonify({"error": "frontend/index.html not found. Please check folder structure."}), 404


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/api/convert", methods=["POST"])
def convert_pdf_to_excel():
    files = request.files.getlist("file")
    if not files or files[0].filename == "":
        return jsonify({"error": "No file uploaded. Field name must be 'file'."}), 400

    valid_files = [f for f in files if f.filename != ""]
    if not valid_files:
        return jsonify({"error": "Empty filename."}), 400

    for f in valid_files:
        fn_lower = f.filename.lower()
        if not (fn_lower.endswith(".pdf") or fn_lower.endswith(IMAGE_EXTENSIONS)):
            return jsonify({"error": "Only .pdf and image files (.png, .jpg, .jpeg, .webp, .bmp) are supported."}), 400

    try:
        first_file = valid_files[0]
        fn_lower = first_file.filename.lower()

        try:
            if fn_lower.endswith(".pdf"):
                file_bytes = first_file.read()
                file_stream = io.BytesIO(file_bytes)
                dataframes = extract_tables_from_pdf(file_stream)
            else:
                dataframes = _extract_tables_from_images(valid_files)
        except Exception as parse_err:
            return jsonify({"error": f"Invalid or corrupted file: {str(parse_err)}"}), 400

        if not dataframes:
            return jsonify({"error": "No tables found in this file."}), 422

        output = _build_styled_excel_file(dataframes)

        ext = os.path.splitext(first_file.filename)[1]
        original_name = first_file.filename[:-len(ext)] if ext else first_file.filename
        download_name = f"{original_name}_converted.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=download_name,
        )

    except Exception as exc:
        return jsonify({"error": f"Conversion failed: {str(exc)}"}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
